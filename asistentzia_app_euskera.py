
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Asistentzia Sortzailea", page_icon="4d8", layout="wide")

# Goiburua eta logotipoa
st.markdown("""
    <div style='display: flex; align-items: center;'>
        <img src='https://upload.wikimedia.org/wikipedia/commons/thumb/3/38/Streamlit_logo.svg/512px-Streamlit_logo.svg.png' width='60'/>
        <h1 style='margin-left: 20px; color: #2c3e50;'>Asistentzia Sortzailea</h1>
    </div>
    <hr style='border: 1px solid #ccc;'>
""", unsafe_allow_html=True)

st.markdown("""
<style>
    .stButton>button {
        background-color: #3498db;
        color: white;
        border-radius: 5px;
        padding: 0.5em 1em;
        font-weight: bold;
    }
    .stDownloadButton>button {
        background-color: #2ecc71;
        color: white;
        border-radius: 5px;
        padding: 0.5em 1em;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

plantilla_file = st.file_uploader("📤 Txantiloia igo (Asistentzia_txantiloia.xlsx)", type="xlsx")
origen_file = st.file_uploader("📤 Jatorrizko fitxategia igo (1B.xlsx)", type="xlsx")

if plantilla_file and origen_file:
    df_origen = pd.read_excel(origen_file, header=None, engine='openpyxl')
    wb = load_workbook(plantilla_file)
    ws_ikasleak = wb['ikasleak']

    start_row = 4
    for i in range(len(df_origen)):
        row = df_origen.iloc[i]
        if pd.notna(row[9]):
            r = start_row + i
            ws_ikasleak.cell(row=r, column=3, value=row[10])
            ws_ikasleak.cell(row=r, column=4, value=row[11])
            ws_ikasleak.cell(row=r, column=5, value=row[12])
            ws_ikasleak.cell(row=r, column=6, value=row[13])
            ws_ikasleak.cell(row=r, column=7, value=row[14])

    modulos_nombres = [
        "Integrazioa",
        "Sist. Pneum+Hidra",
        "Sist Elektr+Elektron",
        "Fabrikazioa",
        "Marraz",
        "IEP I",
        "Digitalizazioa"
    ]

    for i, nuevo_nombre in enumerate(modulos_nombres):
        hoja_original = f"{i+1}.modulua"
        ws = wb[hoja_original]
        ws.title = nuevo_nombre
        ws["M2"] = nuevo_nombre

    dias = ["A", "Astelehena", "Asteartea", "Asteazkena", "Osteguna", "Ostirala"]
    dias = ["L", "M", "X", "J", "V"]
    horas_por_modulo = {}
    for nombre_modulo in modulos_nombres:
        st.subheader(f"⏱️ Orduak moduluarentzat: {nombre_modulo}")
        horas = []
        cols = st.columns(5)
        for idx, dia in enumerate(dias):
            horas.append(cols[idx].number_input(f"{dia}", min_value=0, max_value=10, value=0, key=f"{nombre_modulo}_{dia}"))
        horas_por_modulo[nombre_modulo] = horas

    def rellenar_bloques(ws, fila, repeticiones, horas):
        for i in range(repeticiones):
            start_col = 4 + i * 5
            for j in range(5):
                ws.cell(row=fila, column=start_col + j, value=horas[j])

    for nombre_modulo in modulos_nombres:
        ws = wb[nombre_modulo]
        horas = horas_por_modulo[nombre_modulo]
        rellenar_bloques(ws, fila=6, repeticiones=11, horas=horas)
        rellenar_bloques(ws, fila=37, repeticiones=14, horas=horas)
        rellenar_bloques(ws, fila=69, repeticiones=13, horas=horas)

    output = BytesIO()
    wb.save(output)
    st.success("✅ Fitxategia ondo sortu da")
    st.download_button(
        label="⬇️ Deskargatu Asistentzia_osatua.xlsx",
        data=output.getvalue(),
        file_name="Asistentzia_osatua.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
