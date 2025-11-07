
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.title("📘 Generador de Asistentzia")

plantilla_file = st.file_uploader("📄 Sube la plantilla Asistentzia_txantiloia.xlsx", type="xlsx")
origen_file = st.file_uploader("📄 Sube el fichero de origen 1B.xlsx", type="xlsx")

if plantilla_file and origen_file:
    df_origen = pd.read_excel(origen_file, header=None, engine='openpyxl')
    wb = load_workbook(plantilla_file)
    ws_ikasleak = wb['ikasleak']

    start_row = 4
    for i in range(len(df_origen)):
        row = df_origen.iloc[i]
        if pd.notna(row[9]):
            r = start_row + i
            ws_ikasleak.cell(row=r, column=3, value=row[10])
            ws_ikasleak.cell(row=r, column=4, value=row[11])
            ws_ikasleak.cell(row=r, column=5, value=row[12])
            ws_ikasleak.cell(row=r, column=6, value=row[13])
            ws_ikasleak.cell(row=r, column=7, value=row[14])

    modulos_nombres = [
        "Integrazioa",
        "Sist. Pneum+Hidra",
        "Sist Elektr+Elektron",
        "Fabrikazioa",
        "Marraz",
        "IEP I",
        "Digitalizazioa"
    ]

    for i, nuevo_nombre in enumerate(modulos_nombres):
        hoja_original = f"{i+1}.modulua"
        ws = wb[hoja_original]
        ws.title = nuevo_nombre
        ws["M2"] = nuevo_nombre

    dias = ["L", "M", "X", "J", "V"]
    horas_por_modulo = {}
    for nombre_modulo in modulos_nombres:
        st.subheader(f"Horas para {nombre_modulo}")
        horas = []
        for dia in dias:
            horas.append(st.number_input(f"{dia} ({nombre_modulo})", min_value=0, max_value=10, value=0, key=f"{nombre_modulo}_{dia}"))
        horas_por_modulo[nombre_modulo] = horas

    def rellenar_bloques(ws, fila, repeticiones, horas):
        for i in range(repeticiones):
            start_col = 4 + i * 5
            for j in range(5):
                ws.cell(row=fila, column=start_col + j, value=horas[j])

    for nombre_modulo in modulos_nombres:
        ws = wb[nombre_modulo]
        horas = horas_por_modulo[nombre_modulo]
        rellenar_bloques(ws, fila=6, repeticiones=11, horas=horas)
        rellenar_bloques(ws, fila=37, repeticiones=14, horas=horas)
        rellenar_bloques(ws, fila=69, repeticiones=13, horas=horas)

    output = BytesIO()
    wb.save(output)
    st.success("✅ Archivo generado correctamente")
    st.download_button(
        label="⬇️ Descargar Asistentzia_osatua.xlsx",
        data=output.getvalue(),
        file_name="Asistentzia_osatua.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
